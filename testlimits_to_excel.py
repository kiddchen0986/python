import os
import json
from pysettings.loader import Loader
from shutil import copy, rmtree

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from collections import OrderedDict, Counter
import pandas as pd
from pathlib import Path
import logging


os.chdir(os.path.dirname(__file__))
script_path = os.getcwd()
os.chdir(os.sep.join(script_path.split(os.sep)[:-3]))
sln_root_path = os.getcwd()

folder_name = 'CustomerLimitsFile'
testlimits_path = os.path.join(sln_root_path, "sensor", "sensor_settings",  folder_name)
output_path = os.path.join(sln_root_path, "production_test_package_capacitive", "out", "desktop", "x86", "Release", folder_name)

def get_mtt_version():
    file_path = os.path.join(os.path.dirname(__file__), "tools", "mtt_version")
    if os.path.exists(Path(file_path).as_posix()):
        with open(file_path, mode="r") as f:
            mtt_version = f.read().strip()
    else:
        mtt_version = "Unknown"
    return "MTT Version: " + mtt_version


def my_border(t_border, b_border, l_border, r_border):
    border = Border(left=Side(border_style=l_border, color='000000'),
                    right=Side(border_style=r_border, color='000000'),
                    top=Side(border_style=t_border, color='000000'),
                    bottom=Side(border_style=b_border, color='000000'))
    return border


def format_border(sheet, s_column, s_index, e_column, e_index):
    for row in tuple(sheet[s_column + str(s_index):e_column + str(e_index)]):
        for cell in row:
            cell.border = my_border('thin', 'thin', 'thin', 'thin')


def format_font(sheet, s_column, s_index, e_column, e_index):
    for row in tuple(sheet[s_column + str(s_index):e_column + str(e_index)]):
        for cell in row:
            cell.font = Font(bold=True)


def format_alignment(sheet, s_column, s_index, e_column, e_index):
    for row in tuple(sheet[s_column + str(s_index):e_column + str(e_index)]):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def test_items_info(ws):
    test_items = []             # get all test types, i.e. all data from 1st column
    test_item_row_number = {}   # get all row number for one kind of test
    test_item_info = {}         # get row's index of 1st displayed, number of row for one kind of test
    for row in range(3, ws.max_row + 1):
        test_items.append(str(ws.cell(row, 1).value))
        if test_item_row_number.get(str(ws.cell(row, 1).value)) is not None:
            current = test_item_row_number[str(ws.cell(row, 1).value)]
            current.append(row)
            test_item_row_number[str(ws.cell(row, 1).value)] = current
        else:
            test_item_row_number[str(ws.cell(row, 1).value)] = [row]

    for item, number in dict(Counter(test_items)).items():
        for test_item, row in test_item_row_number.items():
            if item == test_item:
                test_item_info[item] = [min(row), number]

    return test_item_info


def ws_template(ws):
    sheet_rows = 11
    # column width
    ws.column_dimensions['A'].width = 25.0
    ws.column_dimensions['B'].width = 35.0
    ws.column_dimensions['C'].width = 30.0
    ws.column_dimensions['D'].width = 30.0
    # cell handle
    # row 1
    ws['A1'] = get_mtt_version()
    ws.merge_cells('A1:D1')
    # row 2
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    # row > 2
    for test_item, info in test_items_info(ws).items():
        # info[0] is index of row, info[1] is number of row for one test_item
        if info[1] > 1:
            ws.merge_cells(start_row=info[0], start_column=1, end_row=(info[0] + info[1]-1), end_column=1)

    # Border and font formal
    format_font(ws, 'A', 1, 'D', 2)
    format_border(ws, 'A', 1, 'D', sheet_rows)
    format_alignment(ws, 'A', 1, 'D', sheet_rows)
    return ws


def copy_file_to_folder():
    if os.path.exists(output_path):
        rmtree(output_path)
    os.makedirs(output_path)

    # retrieve non-mobile sensors
    non_mobile_sensor_list = []
    try:
        with open("tools/non_mobile_sensors.json",'r') as f:
            load_dict = json.load(f)
            for sensor in load_dict['sensor_type']:
                non_mobile_sensor_list.append(sensor)
    except Exception as e:
        print(e)

    for file in os.listdir(testlimits_path):
        # Only filter out mobile's sensors(FPC1291 and later) supported in repo
        if int(str(file)[3:7]) > 1290:
            if file.split("_")[0] not in non_mobile_sensor_list:   # exclude non-mobile sensors
                print("Copying file: {}".format(os.path.join(testlimits_path, file)))
                copy(os.path.join(testlimits_path, file), output_path)


def gradient_checkerboard_limits(test_name, limit_data):
    try:
        for test_type in limit_data.keys():
            ordered_dict = OrderedDict()
            max_deviation = limit_data['maxDeviation']
            max_pixel_errors = limit_data['pixelErrorsUpperLimit']
            median_delta_Limit = limit_data['medianDeltaLimit']
            if type(limit_data[test_type]) is dict:
                type1_min_value = limit_data[test_type]["type1"]['min']
                type1_max_value = limit_data[test_type]["type1"]['max']
                type2_min_value = limit_data[test_type]["type2"]['min']
                type2_max_value = limit_data[test_type]["type2"]['max']

                # MTS test limits
                mts_value = "type1: (" + str(type1_min_value) + ",  " + str(type1_max_value) + ") \r\n"
                mts_value += "type2: (" + str(type2_min_value) + ",  " + str(type2_max_value) + ")\r\n"
                mts_value += "max Deviation: [" + str(-max_deviation) + ", +" + str(max_deviation) + "]\r\n"
                mts_value += "median Delta Limit: > " + str(median_delta_Limit) + "\r\n"
                mts_value += "max pixel errors: < " + str(max_pixel_errors)

                # MTS-I test limits
                mts_i_value = "type1: (" + str(type1_min_value - 5) + ",  " + str(
                    type1_max_value + 5) + ") \r\n"
                mts_i_value += "type2: (" + str(type2_min_value - 5) + ",  " + str(
                    type2_max_value + 5) + ")\r\n"
                mts_i_value += "max Deviation: [" + str(-max_deviation - 10) + ", +" + str(max_deviation + 10) + "]\r\n"
                mts_i_value += "median Delta Limit: > " + str(median_delta_Limit - 5) + "\r\n"
                mts_i_value += "max pixel errors:  < " + str(max_pixel_errors + 5)
            else:
                continue

            ordered_dict["Test Item"] = "CTL Defective Pixels"
            ordered_dict["Test Sub-Item"] = test_type
            ordered_dict['OQC(MTS)'] = mts_value
            ordered_dict['IQC(MTS-I)'] = mts_i_value
            yield ordered_dict

    except KeyError as e:
        logging.error(
            'KeyError {} happens in Test {}, might be no test result, regard all nan'.format(e, test_name))


def image_constant_limits(test_name, limit_data):
    try:
        ordered_dict = OrderedDict()
        # MTS
        mts_value = "(" + str(limit_data['medianLimits']['min']) + ", " + str(
            limit_data['medianLimits']['max']) + ")\r\n"
        mts_value += "max Deviation: [" + str(-limit_data["maxMedianDeviation"]) + ", +" + str(
            limit_data["maxMedianDeviation"]) + "]\r\n"
        mts_value += "max pixel errors: < " + str(limit_data["pixelErrorsUpperLimit"])
        # MTS-I
        mts_i_value = "(" + str(limit_data['medianLimits']['min'] - 5) + ", " + str(
            limit_data['medianLimits']['max'] + 5) + ")\r\n"
        mts_i_value += "max Deviation: [" + str(-limit_data["maxMedianDeviation"] - 10) + ", +" + str(
            limit_data["maxMedianDeviation"] + 10) + "]\r\n"
        mts_i_value += "max pixel errors: < " + str(limit_data["pixelErrorsUpperLimit"] + 5)

        ordered_dict["Test Item"] = "CTL Defective Pixels"
        ordered_dict["Test Sub-Item"] = test_name
        ordered_dict['OQC(MTS)'] = mts_value
        ordered_dict['IQC(MTS-I)'] = mts_i_value
        yield ordered_dict

    except KeyError as e:
        logging.error(
            'KeyError {} happens in Test {}, might be no test result, regard all nan'.format(e, test_name))


def current_consumption_limits(test_name, limit_data):
    for test_type in limit_data.keys():
        try:
            if type(limit_data[test_type]) is dict:
                ordered_dict = OrderedDict()
                delta_limit = 0.5
                rail = "（" + limit_data["rails"] + "）"
                if test_type == 'deep_sleep':
                    # MTS
                    mts_value = "(" + str(float(limit_data["deep_sleep"]["low"] * 1000)) + "uA, "
                    mts_value += str(float(limit_data["deep_sleep"]["high"] * 1000)) + "uA)"
                    # MTS-I
                    mts_i_value = "(" + str(float(limit_data["deep_sleep"]["low"] * 1000) - delta_limit) + "uA, "
                    mts_i_value += str(float(limit_data["deep_sleep"]["high"] * 1000) + delta_limit) + "uA)"

                    ordered_dict["Test Item"] = "Current Consumption"
                    ordered_dict["Test Sub-Item"] = test_type.title() + rail
                    ordered_dict['OQC(MTS)'] = mts_value
                    ordered_dict['IQC(MTS-I)'] = mts_i_value
                if test_type == 'active_hp':
                    # MTS
                    mts_value = "(" + str(float(limit_data["active_hp"]["low"])) + "mA, "
                    mts_value += str(float(limit_data["active_hp"]["high"])) + "mA)"
                    # MTS-I
                    mts_i_value = "(" + str(float(limit_data["active_hp"]["low"]) - delta_limit) + "mA, "
                    mts_i_value += str(float(limit_data["active_hp"]["high"]) + delta_limit) + "mA)"

                    ordered_dict["Test Item"] = "Current Consumption"
                    ordered_dict["Test Sub-Item"] = test_type.title() + rail
                    ordered_dict['OQC(MTS)'] = mts_value
                    ordered_dict['IQC(MTS-I)'] = mts_i_value
                yield ordered_dict

        except KeyError as e:
            logging.error(
                'KeyError {} happens in Test {}, might be no test result, regard all nan'.format(e, test_name))

def module_quality2_limits(test_name, limit_data):
    try:
        for test_type in limit_data.keys():
            if type(limit_data[test_type]) is dict and limit_data[test_type].get("enabled"):
                ordered_dict = OrderedDict()
                if test_type == "snrTest":
                    # MTS
                    mts_value = "> " + str(limit_data[test_type]['snrLimit'])
                    # MTS-I
                    mts_i_value = "> " + str(limit_data[test_type]['snrLimit'] - 2)
                elif test_type == "udrTest":
                    # MTS
                    mts_value = "> " + str(limit_data[test_type]["udrLimit"])
                    # MTS-I same as MTS
                    mts_i_value = mts_value
                elif test_type == "fixedPatternTest":
                    # MTS
                    mts_value = "< " + str(limit_data[test_type]["fixedPatternLimit"])
                    # MTS-I
                    mts_i_value = "< " + str(limit_data[test_type]["fixedPatternLimit"] + 20)
                else:
                    continue

                ordered_dict["Test Item"] = "Module Quality Test 2"
                ordered_dict["Test Sub-Item"] = test_type
                ordered_dict['OQC(MTS)'] = mts_value
                ordered_dict['IQC(MTS-I)'] = mts_i_value
                yield ordered_dict

    except KeyError as e:
        logging.error(
            'KeyError {} happens in Test {}, might be no test result, regard all nan'.format(e, test_name))


def generate_test_limits_file(test_limits, data_list):
    for test_name in test_limits.get_test_names():
        limit_data = test_limits.get_limits_test_data(test_name)
        if test_name == "gradientCheckerboard":
            data_list.extend(gradient_checkerboard_limits(test_name, limit_data))

        if test_name == "imageConstant":
            data_list.extend(image_constant_limits(test_name, limit_data))

        if test_name == "currentConsumption":
            data_list.extend(current_consumption_limits(test_name, limit_data))

        if test_name == "moduleQuality2":
            data_list.extend(module_quality2_limits(test_name, limit_data))


def format_xlsx_template(xlsx):
    wb = load_workbook(xlsx)
    ws = wb.active
    ws_template(ws)
    wb.save(xlsx)


def process_and_update(settings):
    try:
        for product in settings.get_products():
            data_list = []
            print("Generating %s test limits to Excel..." % (product.get_product_name()))

            test_limits = product.get_limits()
            generate_test_limits_file(test_limits, data_list)

            data_frame = pd.DataFrame(data_list).dropna(axis=0)
            data_frame.to_excel(os.path.join(testlimits_path, '%s.xlsx' % (product.get_product_name().replace("?", ""))),
                                index=False, startrow=1, encoding='utf-8')

            format_xlsx_template(os.path.join(testlimits_path, '%s.xlsx' % (product.get_product_name().replace("?", ""))))

    except FileNotFoundError as e:
        print("Error: File {} Not Found".format(e))
    except Exception as e:
        print("raise error {}".format(e))

    print("Finished to generate test limits file")
    print("*************************************")


def main():
    loader = Loader('.')
    config = loader.load_config("config.json")
    manifest = loader.load_manifest("tools/manifest_testdata.json")
    for source in manifest.get_sources():
        print('Processing %s' % source.get_name())
        settings = loader.load_settings(source, config)
        process_and_update(settings)

    print("Copy all limits file to {}...".format(output_path))
    copy_file_to_folder()

    print("Done")


if __name__ == "__main__":

    if not os.path.exists(os.path.join(sln_root_path, "production_test_package_capacitive")):
        print("This is not MTT capacitive tool. This script will exit because don't need to generate testlimits file!")
        exit(0)

    if not os.path.exists(testlimits_path):
        os.makedirs(testlimits_path)

    os.chdir(script_path)
    main()
